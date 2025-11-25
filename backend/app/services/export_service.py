"""
Serviço de Exportação
======================

Gerencia exportação de dados de planos de intervenção para CSV e Excel.

Autor: Claude Code
Data: 2025-11-24
"""

import csv
import io
import logging
from datetime import datetime
from typing import List, Optional
from uuid import UUID

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None

from sqlalchemy.orm import Session

from app.models.intervention_plan import InterventionPlan
from app.services.intervention_plan_service import InterventionPlanService

logger = logging.getLogger(__name__)


class ExportService:
    """Serviço para exportação de dados."""

    def __init__(self, db: Session):
        self.db = db
        self.plan_service = InterventionPlanService(db)

    def _get_plan_row_data(self, plan: InterventionPlan, include_student: bool = False) -> dict:
        """
        Extrai dados de um plano para exportação.

        Args:
            plan: Plano de intervenção
            include_student: Se True, inclui dados do aluno

        Returns:
            Dicionário com dados do plano
        """
        row = {
            "ID": str(plan.id),
            "Título": plan.title,
            "Descrição": plan.description or "",
            "Status": plan.status.value,
            "Frequência de Revisão": plan.review_frequency.value,
            "Precisa Revisão": "Sim" if plan.needs_review else "Não",
            "Última Revisão": (
                plan.last_reviewed_at.strftime("%d/%m/%Y") if plan.last_reviewed_at else "Nunca"
            ),
            "Data de Criação": plan.created_at.strftime("%d/%m/%Y %H:%M"),
            "Data de Atualização": (
                plan.updated_at.strftime("%d/%m/%Y %H:%M") if plan.updated_at else ""
            ),
        }

        if include_student and plan.student:
            row.update(
                {
                    "Aluno ID": str(plan.student_id),
                    "Aluno Nome": plan.student.name or "Não informado",
                    "Aluno Idade": plan.student.age if hasattr(plan.student, "age") else "",
                }
            )

        return row

    def export_to_csv(
        self,
        skip: int = 0,
        limit: int = 1000,
        priority_filter: Optional[str] = None,
        professional_id: Optional[UUID] = None,
        include_student: bool = False,
    ) -> str:
        """
        Exporta planos pendentes de revisão para CSV.

        Args:
            skip: Offset para paginação
            limit: Limite de registros (máximo 1000)
            priority_filter: Filtro de prioridade
            professional_id: ID do profissional
            include_student: Incluir dados do aluno

        Returns:
            String CSV
        """
        logger.info(
            "Exporting pending review plans to CSV",
            extra={
                "skip": skip,
                "limit": min(limit, 1000),
                "priority_filter": priority_filter,
                "professional_id": str(professional_id) if professional_id else None,
            },
        )

        # Obter planos
        result = self.plan_service.get_pending_review_plans(
            skip=skip, limit=min(limit, 1000), priority_filter=priority_filter, professional_id=professional_id
        )

        plans = result["items"]

        if not plans:
            logger.warning("No plans found for export")
            return ""

        # Preparar CSV
        output = io.StringIO()

        # Determinar colunas
        first_row = self._get_plan_row_data(plans[0]["plan"], include_student=include_student)
        fieldnames = list(first_row.keys())

        # Adicionar coluna de prioridade
        fieldnames.insert(1, "Prioridade")

        writer = csv.DictWriter(output, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()

        # Escrever linhas
        for item in plans:
            plan = item["plan"]
            priority = item["priority"]

            row = self._get_plan_row_data(plan, include_student=include_student)
            row["Prioridade"] = priority.upper()

            writer.writerow(row)

        logger.info(f"Exported {len(plans)} plans to CSV")
        return output.getvalue()

    def export_to_excel(
        self,
        skip: int = 0,
        limit: int = 1000,
        priority_filter: Optional[str] = None,
        professional_id: Optional[UUID] = None,
        include_student: bool = False,
    ) -> Optional[bytes]:
        """
        Exporta planos pendentes de revisão para Excel.

        Args:
            skip: Offset para paginação
            limit: Limite de registros (máximo 1000)
            priority_filter: Filtro de prioridade
            professional_id: ID do profissional
            include_student: Incluir dados do aluno

        Returns:
            Bytes do arquivo Excel ou None se openpyxl não disponível
        """
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl not available, cannot export to Excel")
            return None

        logger.info(
            "Exporting pending review plans to Excel",
            extra={
                "skip": skip,
                "limit": min(limit, 1000),
                "priority_filter": priority_filter,
                "professional_id": str(professional_id) if professional_id else None,
            },
        )

        # Obter planos
        result = self.plan_service.get_pending_review_plans(
            skip=skip, limit=min(limit, 1000), priority_filter=priority_filter, professional_id=professional_id
        )

        plans = result["items"]

        if not plans:
            logger.warning("No plans found for export")
            return None

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planos Pendentes de Revisão"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Determinar colunas
        first_row = self._get_plan_row_data(plans[0]["plan"], include_student=include_student)
        headers = ["Prioridade"] + list(first_row.keys())

        # Escrever cabeçalhos
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Escrever dados
        for row_idx, item in enumerate(plans, start=2):
            plan = item["plan"]
            priority = item["priority"]

            row_data = self._get_plan_row_data(plan, include_student=include_student)
            values = [priority.upper()] + list(row_data.values())

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Colorir por prioridade
                if col_idx == 1:  # Coluna de prioridade
                    if priority == "high":
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )
                    elif priority == "medium":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )

        # Ajustar largura das colunas
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20

        # Adicionar sumário em nova aba
        ws_summary = wb.create_sheet("Resumo")
        ws_summary["A1"] = "Resumo da Exportação"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "Total de Planos:"
        ws_summary["B3"] = result["total"]

        ws_summary["A4"] = "Alta Prioridade:"
        ws_summary["B4"] = result["high_priority"]

        ws_summary["A5"] = "Média Prioridade:"
        ws_summary["B5"] = result["medium_priority"]

        ws_summary["A6"] = "Baixa Prioridade:"
        ws_summary["B6"] = result["low_priority"]

        ws_summary["A8"] = "Data da Exportação:"
        ws_summary["B8"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Ajustar largura
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15

        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Exported {len(plans)} plans to Excel")
        return output.getvalue()

    def get_export_summary(
        self, priority_filter: Optional[str] = None, professional_id: Optional[UUID] = None
    ) -> dict:
        """
        Retorna resumo dos dados que serão exportados.

        Args:
            priority_filter: Filtro de prioridade
            professional_id: ID do profissional

        Returns:
            Dicionário com resumo
        """
        result = self.plan_service.get_pending_review_plans(
            skip=0, limit=1, priority_filter=priority_filter, professional_id=professional_id
        )

        return {
            "total": result["total"],
            "high_priority": result["high_priority"],
            "medium_priority": result["medium_priority"],
            "low_priority": result["low_priority"],
            "excel_available": EXCEL_AVAILABLE,
        }
